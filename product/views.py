import json

from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView
from django.db.models.functions import ExtractYear
from .serializers import *
from rest_framework import generics
from rest_framework import generics, viewsets, parsers
import django_filters
from django.db.models import Q
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
import django_filters
from rest_framework import filters
import shutil
from django.conf import settings

class ProductPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 10000

    # def get_paginated_response(self, data):
    #     print(self.to_html())
    #     return Response({
    #         'links':{
    #             'next': self.get_next_link(),
    #             'prev': self.get_previous_link(),
    #         },
    #         'page_count':self.page.paginator.num_pages,
    #         'results':data
    #     })

class CategoryViewSet(viewsets.ModelViewSet):
    serializer_class = ProductCategorySerializer
    queryset = ProductCategory.objects.all()
    lookup_field = 'slug'

class FilterViewSet(viewsets.ModelViewSet):
    serializer_class = FilterSerializer
    queryset = Filter.objects.all()
    lookup_field = 'slug'


class ProductPriceFilter(django_filters.FilterSet):
    q = django_filters.CharFilter(method='my_custom_filter', label="Search")
    def my_custom_filter(self, queryset, name, value):
        return queryset.filter(
            Q(product__name__icontains=value) |
            Q(textLabel__icontains=value) |
            Q(vendorCode__icontains=value)
        )
    class Meta:
        model = ProductPrice
        fields = ['id']




class GetProductPrices(generics.ListAPIView):
    serializer_class = ProductPriceForTableSerializer
    queryset = ProductPrice.objects.all()
    pagination_class = ProductPagination
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = ProductPriceFilter
    ordering_fields = ['weight', 'price', 'at_store','isActive']


class GetProductsByFilter(generics.ListAPIView):
    serializer_class = ProductShortSerializer
    def get_queryset(self):
        return Product.objects.filter(category__slug=self.kwargs['cat_slug'],filters__slug__in=[self.kwargs['filter_slug']])


class ProductFilter(django_filters.FilterSet):
    q = django_filters.CharFilter(method='my_custom_filter', label="Search")
    def my_custom_filter(self, queryset, name, value):
        return queryset.filter(
            Q(name__icontains=value) |
            Q(vendorCode__icontains=value)
        )
    class Meta:
        model = Product
        fields = ['id','category__slug']

class ProductPriceViewSet(viewsets.ModelViewSet):
    serializer_class = ProductPriceSerializer
    queryset = ProductPrice.objects.all()

class ProductTabViewSet(viewsets.ModelViewSet):
    serializer_class = ProductTabSerializer
    queryset = ProductTab.objects.all()

class ProductImagesViewSet(viewsets.ModelViewSet):
    serializer_class = ProductGalleryImageSerializer
    queryset = ProductGalleryImage.objects.all()





class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    queryset = Product.objects.all().order_by('id')
    pagination_class = ProductPagination
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]

    filterset_class = ProductFilter


    def update(self, request, *args, **kwargs):
        images = request.FILES.getlist('images', None)
        print(images)
        data = json.loads(request.data['data'])
        main_image = request.data['main_img']
        print(main_image)
        print(data)
        id = data['id']
        product = Product.objects.get(id=id)
        name = data['name']
        isNew = data['isNew']
        isAvailable = data['isAvailable']
        vendorCode = data['vendorCode']
        shortDescription = data['shortDescription']
        filters = data['filters']
        prices = data['prices']
        tabs = data['tabs']
        category_id = data['category']['id']

        product.category_id = category_id
        product.name = name
        product.isNew = isNew
        product.isAvailable = isAvailable
        product.vendorCode = vendorCode
        product.shortDescription = shortDescription
        product.save()
        product.filters.clear()
        for filter in filters:
            product.filters.add(filter['id'])

        for tab in tabs:
            is_new = tab.get('is_new',None)
            print(tab)
            if is_new:
                ProductTab.objects.create(
                    product=product,
                    label=tab['label'],
                    text=tab['text']
                )
            else:
                tab_obj = ProductTab.objects.get(id=tab['id'])
                tab_obj.label = tab['label']
                tab_obj.text = tab['text']
                tab_obj.save()

        for price in prices:
            is_new = price.get('is_new', None)
            if is_new:
                ProductPrice.objects.create(
                    product=product,
                    vendorCode=price['vendorCode'],
                    weight=price['weight'],
                    volume=price['volume'],
                    textLabel=price['textLabel'],
                    price=price['price'],
                    at_store=price['at_store'],
                )
            else:
                price_obj = ProductPrice.objects.get(id=price['id'])
                price_obj.vendorCode = price['vendorCode']
                price_obj.weight = price['weight']
                price_obj.volume = price['volume']
                price_obj.textLabel = price['textLabel']
                price_obj.price = price['price']
                price_obj.at_store = price['at_store']
                price_obj.save()

        for image in images:
            ProductGalleryImage.objects.create(
                product=product,
                image=image,
                is_main=False
            )


        return Response({'success': True}, status=200)
    def create(self, request, *args, **kwargs):
        try:
            images = request.FILES.getlist('images',None)
            print(images)
            data = json.loads(request.data['data'])
            main_image = request.data['main_img']
            print(main_image)
            name = data['name']
            isNew = data['isNew']
            isAvailable = data['isAvailable']
            vendorCode = data['vendorCode']
            shortDescription = data['shortDescription']
            filters = data['filters']
            prices = data['prices']
            tabs = data['tabs']
            category_id = data['category']['id']
            new_product = Product.objects.create(
                category_id=category_id,
                name=name,
                isNew=isNew,
                isAvailable=isAvailable,
                vendorCode=vendorCode,
                shortDescription=shortDescription,
            )
            for filter in filters:
                new_product.filters.add(filter['id'])
            for tab in tabs:
                ProductTab.objects.create(
                    product=new_product,
                    label=tab['label'],
                    text=tab['text']
                )
            for price in prices:
                ProductPrice.objects.create(
                    product=new_product,
                    vendorCode=price['vendorCode'],
                    weight=price['weight'],
                    volume=price['volume'],
                    textLabel=price['textLabel'],
                    price=price['price'],
                    at_store=price['at_store'],

                )
            i = 0
            for image in images:
                ProductGalleryImage.objects.create(
                    product=new_product,
                    image=image,
                    is_main= True if i == int(main_image) else False
                )
                i += 1

            return Response({'success':True},status=200)
        except:
            return Response(status=500)





class FillCat(APIView):
    def get(self,request):
        from openpyxl import load_workbook
        cats = ProductCategory.objects.all()
        cats.delete()
        wb = load_workbook(filename='category.xlsx')
        sheet_obj = wb.active
        max_row = sheet_obj.max_row

        # Loop will print all columns name max_row + 1

        for i in range(2, max_row + 1):
            name = sheet_obj.cell(row=i, column=1)
            old_id = sheet_obj.cell(row=i, column=2)
            print(name.value,old_id.value)
            ProductCategory.objects.create(
                name=name.value,
                old_id=old_id.value
            )
        return Response(status=200)


class FillProd(APIView):
    def get(self,request):
        from openpyxl import load_workbook
        prods = Product.objects.all()
        prods.delete()
        wb = load_workbook(filename='product.xlsx')
        sheet_obj = wb.active
        max_row = sheet_obj.max_row

        # Loop will print all columns name max_row + 1

        for i in range(2, max_row + 1):
            cat_id = sheet_obj.cell(row=i, column=1)
            name = sheet_obj.cell(row=i, column=2)
            descr = sheet_obj.cell(row=i, column=3)
            code = sheet_obj.cell(row=i, column=4)

            print(cat_id.value,name.value,descr.value,code.value)


            if cat_id.value:
                cat = ProductCategory.objects.get(old_id=cat_id.value)
                Product.objects.create(
                    category=cat,
                    name=name.value,
                    shortDescription=descr.value,
                    vendorCode=code.value
                )

        return Response(status=200)

class FillFas(APIView):
    def get(self,request):
        from openpyxl import load_workbook
        fas = ProductPrice.objects.all()
        fas.delete()
        wb = load_workbook(filename='fasovka.xlsx')
        sheet_obj = wb.active
        max_row = sheet_obj.max_row

        # Loop will print all columns name max_row + 1

        for i in range(2, max_row + 1):
            product_code = sheet_obj.cell(row=i, column=1)
            name = sheet_obj.cell(row=i, column=2)
            code = sheet_obj.cell(row=i, column=3)
            price = sheet_obj.cell(row=i, column=3)
            print(product_code.value, name.value, code.value)
            prod = Product.objects.get(vendorCode=product_code.value)
            ProductPrice.objects.create(
                product=prod,
                vendorCode=code.value,
                textLabel=name.value,
                price=price.value if price.value !='-' else 0

            )



        return Response(status=200)