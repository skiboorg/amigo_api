import json

import django_filters
from django.db.models import Q
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from .serializers import *
from .models import *
from rest_framework import generics, viewsets, parsers
from rest_framework import filters

import logging
logger = logging.getLogger(__name__)



class GetUser(generics.RetrieveAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer

    def get_object(self):
        return self.request.user


class AddUser(APIView):
    def post(self,request):
        print(request.data)
        data = request.data
        result = {}

        if data['password1'] != data['password2']:
            result = {'success':False,'message':'Пароли не совпадают'}
            return Response(result,status=200)
        new_user = User.objects.create(
            email=data['email'],
            client_id=data['client'],
            login=data['email'],
            fio=data['fio'],
            comment=data['comment'],
            is_manager=data['is_manager'],
            is_staff=data['is_staff'],
            plain_password=data['password1'],
        )
        new_user.set_password(data['password1'])
        new_user.save()
        result = {'success': True, 'message': 'Пользователь успешно создан'}
        return Response(result, status=200)



class UpdateUser(APIView):
    def post(self,request,*args,**kwargs):
        print(request.data)
        data = request.data
        password1 = data.get('password1',None)
        user = User.objects.get(id=data['id'])
        if password1:
            if data['password1'] != data['password2']:
                result = {'success': False, 'message': 'Пароли не совпадают'}
                return Response(result, status=200)
            else:
                user.set_password(data['password1'])
                user.plain_password = data['password1']
        user.client_id = data['client']
        user.email = data['email']
        user.login = data['login']
        user.fio = data['fio']
        user.comment = data['comment']
        user.is_manager = data['is_manager']
        user.is_staff = data['is_staff']
        user.save()
        result = {'success': True, 'message': 'Пользователь успешно обновлен'}
        return Response(result,status=status.HTTP_201_CREATED)

class GetMyUsers(generics.ListAPIView):
    serializer_class = UserSerializer

    def get_queryset(self):
        user = User.objects.get(uuid=self.request.query_params.get('id'))
        return User.objects.filter(added_by=user)


class GetUserByID(generics.RetrieveAPIView):
    queryset = User.objects.all()
    serializer_class = UserFullSerializer
    lookup_field = 'id'



class DeleteUser(generics.DestroyAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    lookup_field = 'uuid'



class UserPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 10000


class UserFilter(django_filters.FilterSet):
    q = django_filters.CharFilter(method='my_custom_filter', label="Search")
    def my_custom_filter(self, queryset, name, value):
        return queryset.filter(
            Q(client__fio__icontains=value) |
            Q(fio__icontains=value)

        )
    class Meta:
        model = User
        fields = ['is_manager','is_staff']

class GetAllUsers(generics.ListAPIView):
    # permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer
    queryset = User.objects.filter(is_active=True)
    pagination_class = UserPagination
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = UserFilter

class GetRoles(generics.ListCreateAPIView):
    # permission_classes = [IsAuthenticated]
    serializer_class = RoleSerializer
    queryset = Role.objects.all()

class GetManagers(generics.ListCreateAPIView):
    # permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer

    def get_queryset(self):
        return User.objects.filter(is_manager=True)

class GetUserByRole(generics.ListAPIView):
    # permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer
    def get_queryset(self):
        return User.objects.filter(role__id = self.request.query_params.get('id'))




class FillUser(APIView):
    def get(self,request):
        from openpyxl import load_workbook
        users = User.objects.filter(is_superuser=False)
        users.delete()
        wb = load_workbook(filename='users.xlsx')
        sheet_obj = wb.active
        max_row = sheet_obj.max_row

        # Loop will print all columns name

        for i in range(2, max_row + 1):

            old_id = sheet_obj.cell(row=i, column=1)
            login = sheet_obj.cell(row=i, column=2)
            fio = sheet_obj.cell(row=i, column=3)
            is_manager = sheet_obj.cell(row=i, column=4)
            is_staff = sheet_obj.cell(row=i, column=5)
            print(old_id.value,login.value,fio.value,is_manager.value,is_staff.value)
            u = User.objects.create(
                old_id=old_id.value,
                login=login.value,
                fio=fio.value,
                is_manager = True if is_manager.value==1 else False,
                is_staff = True if is_staff.value==1 else False,
                plain_password=login.value
            )
            u.set_password(login.value)
            u.save()


        return Response(status=200)