
import json
import simplejson
import time
from decimal import Decimal
from rest_framework import viewsets
from rest_framework.pagination import PageNumberPagination
from selenium import webdriver
from selenium.webdriver.common.by import By



from rest_framework.response import Response
from rest_framework.views import APIView
from cart.models import Cart
from cart.views import calcCart
from .models import *
from user.models import User
from client.models import Client,Contractor,Contact
from product.services import create_random_string
from .serializers import *
from django_filters import IsoDateTimeFilter
import django_filters
from django.db.models import Q
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters

import requests
import base64


def calcOrder(order_id):
    order = Order.objects.get(id=order_id)
    total_price = Decimal(0)
    total_weight = Decimal(0)
    for product in order.products.all():
        print(product)
        product.total_price = 0
        product.total_price+= product.price_with_discount * product.amount
        total_price+= product.total_price
        product.total_weight = 0
        product.total_weight += product.productPrice.weight * product.amount
        total_weight += product.total_weight
        product.save()
    order.total_price = total_price
    order.total_weight = total_weight
    order.save()

class OrderItemViewSet(viewsets.ModelViewSet):

    serializer_class = OrderItem
    queryset = OrderItem.objects.all()

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        order_id = instance.order.id
        self.perform_destroy(instance)
        calcOrder(order_id)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        print(obj.price_with_discount)
        print(request.data['manager'])
        print(request.data)
        obj.amount = request.data['amount']
        if Decimal(obj.price_with_discount) != Decimal(request.data['price_with_discount']):
            print('price change')
            obj.price_changed = True
            obj.price_changed_by_id = request.data['manager']
        obj.price_with_discount = request.data['price_with_discount']
        obj.save()
        calcOrder(request.data['order'])
        return Response(status=200)

class OrderPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 10000



class OrderFilter(django_filters.FilterSet):
    q = django_filters.CharFilter(method='my_custom_filter', label="Search")
    created_at_gte = IsoDateTimeFilter(field_name="created_at_date", lookup_expr='gte')
    created_at_lte = IsoDateTimeFilter(field_name="created_at_date", lookup_expr='lte')
    total_price_gte = django_filters.NumberFilter(field_name="total_price", lookup_expr='gte')
    total_price_lte = django_filters.NumberFilter(field_name="total_price", lookup_expr='lte')
    def my_custom_filter(self, queryset, name, value):
        return queryset.filter(
            Q(client__fio__icontains=value) |
            Q(id__exact=value)

        )
    class Meta:
        model = Order
        fields = ['status__id','manager__id','payment_type__id']

class OrderViewSet(viewsets.ModelViewSet):
    pagination_class = OrderPagination
    serializer_class = OrderSerializer
    queryset = Order.objects.all()
    filterset_class = OrderFilter

    def get_queryset(self):
        manager_id = self.request.query_params.get('manager_id',None)

        if manager_id:
            manager = User.objects.get(id=manager_id)
            if manager.is_staff:
                return Order.objects.all()
            else:
                return Order.objects.filter(manager=manager)
        else:
            return Order.objects.all()

    def get_serializer_class(self):
        is_edit = self.request.query_params.get('edit',None)
        if is_edit:
            return OrderEditSerializer
        else:
            return OrderSerializer

    # def perform_create(self, serializer):
    #     print(serializer)

    def create(self, request, *args, **kwargs):
        data = request.data
        #print(data)
        new_order = Order.objects.create(
            client_id=data.get('client',None),
            contact_id=data.get('contact',None),
            manager_id=data.get('manager',None),
            status_id=data.get('status',None),
            contractor_id=data.get('contractor',None),
        )
        need_new_contact = data.get('need_new_contact',None)
        is_admin_order = data.get('is_admin_order',None)

        if not is_admin_order:
            default_status = Status.objects.get(is_default=True)
            new_order.status = default_status
            new_order.save()

        for product in data['products']:
            print(product)
            OrderItem.objects.create(
                order=new_order,
                product_id=product['productPrice']['product'],
                productPrice_id=product['productPrice']['id'],
                price_with_discount=product['price_with_discount'],
                amount=product['amount']
            )
        if need_new_contact:
            contact_data = data['new_contact']
            new_contact = Contact.objects.create(
                client=new_order.client,
                name=contact_data['name'],
            phone=contact_data['phone'],
            email=contact_data['email'],
            invite=contact_data.get('invite',None),
            comment=contact_data.get('comment',None),
            )
            new_order.contact = new_contact
            new_order.save()

        return Response(status=200)

    def update(self, request, *args, **kwargs):
        data = request.data
        print(data)
        instance = self.get_object()
        client_id = data.get('client',None)
        status_id = data.get('status',None)
        contact_id = data.get('contact',None)
        contractor_id = data.get('contractor',None)
        manager_id = data.get('manager',None)
        delivery_id = data.get('delivery',None)
        delivery_company_id = data.get('delivery_company',None)
        delivery_status_id = data.get('delivery_status',None)
        payment_type_id = data.get('payment_type',None)
        is_free_delivery = data.get('is_free_delivery', None)
        delivery_pay_at_office = data.get('delivery_pay_at_office', None)

        if client_id:
            instance.client_id = client_id
        if contact_id:
            instance.contact_id = contact_id
        if contractor_id:
            instance.contractor_id = contractor_id
        if manager_id:
            instance.manager_id = manager_id
        if status_id:
            instance.status_id = status_id
        if delivery_id:
            instance.delivery_id = delivery_id
        if delivery_company_id:
            instance.delivery_company_id = delivery_company_id
        if delivery_status_id:
            instance.delivery_status_id = delivery_status_id
        if payment_type_id:
            instance.payment_type_id = payment_type_id
        instance.track_code = data.get('track_code',None)
        instance.delivery_price = data.get('delivery_price',None)
        instance.delivery_address = data.get('delivery_address',None)
        instance.delivery_comment = data.get('delivery_comment',None)
        instance.is_free_delivery = is_free_delivery
        instance.delivery_pay_at_office = delivery_pay_at_office
        instance.save()
        products = data.get('products', None)
        if products:
            for product in products:
                print(product)
                OrderItem.objects.create(
                    order=instance,
                    product_id=product['productPrice']['product'],
                    productPrice_id=product['productPrice']['id'],
                    price_with_discount=product['price_with_discount'],
                    amount=product['amount']
                )
            calcOrder(instance.id)
        return Response(status=200)

class UpdateOrderItem(APIView):
    def post(self, request):
        data = request.data
        print(data)
        return Response(status=200)


class CreateOrder(APIView):
    def post(self,request):
        data = request.data
        session_id = data['session_id']
        new_user = data.get('new_user', None)
        client = None
        contact = None

        if request.user.is_authenticated:
            cart = Cart.objects.get(user=request.user)
        else:
            cart = Cart.objects.get(sessionID=session_id)

        if new_user:
            email = data['email']
            fio = data['customer_name']
            phone = data['phone']
            client = Client.objects.create(
                fio=fio
            )
            contact = Contact.objects.create(
                client=client,
                phone=phone,
                email=email
            )
            password = create_random_string(digits=False, num=8)
            print(password)
            user = User.objects.create_user(
                client=client,
                login=email,
                email=email,
                password=password,
                plain_password=password
            )
            cart.user = user
            cart.save()

        order = Order.objects.create(
            client=client,
            contact=contact
        )

        order.totalPrice = cart.totalPrice
        order.orderComment = data['order_comment']
        order.delivery_address = data['delivery_address']
        order.delivery_comment = data['delivery_comment']
        order.delivery_id = data['delivery_type_id']
        order.payment_type_id = data['payment_type_id']

        order.save()
        for cart_product in cart.products.all():
            OrderItem.objects.create(
                order=order,
                product=cart_product.product,
                productPrice=cart_product.productPrice,
                amount=cart_product.amount,
                total_price=cart_product.totalPrice,

            )
            cart_product.delete()
            calcCart(cart)

        return Response(status=200)

class DeliveryViewSet(viewsets.ModelViewSet):
    queryset = Delivery.objects.all()
    serializer_class = DeliverySerializer



class PaymentTypeViewSet(viewsets.ModelViewSet):
    queryset = PaymentType.objects.all()
    serializer_class = PaymentTypeSerializer

class StatusViewSet(viewsets.ModelViewSet):
    queryset = Status.objects.all()
    serializer_class = StatusSerializer
def get_from_table(table):
    rows = table.find_elements(By.TAG_NAME,'tr')

    # Создать список для хранения данных
    data_list = []

    # Проход по каждой строке и извлечение данных из ячеек (td)
    for row in rows:
        # Получить все ячейки (td) из текущей строки
        cells = row.find_elements(By.TAG_NAME,'td')
        row_data = []

        # Проход по каждой ячейке и добавление ее содержимого в список данных текущей строки
        for cell in cells:
            row_data.append(cell.text)

        # Добавление списка данных текущей строки в общий список данных
        data_list.append(row_data)

    # Вывод результатов
    # for row_data in data_list:
    #     print(row_data)
    return data_list



def encode_login_password(login,password):
    message = f'{login}:{password}'
    message_bytes = message.encode('ascii')
    base64_bytes = base64.b64encode(message_bytes)
    base64_message = base64_bytes.decode('ascii')
    return base64_message

def get_paykeeper_token(server,login,password):
    uri = f"{server}/info/settings/token/"

    headers = {
        'Content-Type': 'application/x-www-form-urlencoded',
        'Authorization': f'Basic {encode_login_password(login,password)}'
    }
    response = requests.get(uri, headers=headers)
    if response.status_code == 200:
        return {"success": True, 'token': response.json()['token']}
    else:
        return {"success": False}


def create_payment_link(auth_token,server,login,password,order):
    uri = f"{server}/change/invoice/preview/"
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded',
        'Authorization': f'Basic {encode_login_password(login, password)}'
    }
    cart = []
    for item in order.products.all():
        cart.append({"name": item.product.name,
                     "price": item.price_with_discount,
                     "quantity": item.amount,
                     "sum": item.price_with_discount * item.amount,
                     "tax": "none",
                     "item_type": "payment",
                     "payment_type": "prepay"}
                    )
    payload = {
        "pay_amount": Decimal(order.total_price),
        "clientid": order.client.fio if order.client else 'ФИО клиента не указано',
        "orderid": order.id,
        "service_name": {},
        "client_email": order.contact.email if order.contact else None,
        "client_phone": order.contact.phone if order.contact else None,
        "token": auth_token,
    }

    payload['service_name'] = simplejson.dumps({
            'cart': cart,
            'service_name': 'Оплата товара'
    })

    response = requests.post(uri, headers=headers, data=payload)
    data = response.json()

    if data.get('result') == 'fail':
        return {"success": False}
    else:
        return {"success": True, 'data': data}


class GetPaymentLink(APIView):
    def post(self, request):
        data=request.data
        result = {}
        order = Order.objects.get(id=data['order_id'])
        payment_type = order.payment_type
        print(payment_type)
        get_token = get_paykeeper_token(payment_type.url,payment_type.login,payment_type.password)
        if get_token['success']:
            auth_token = get_token['token']
            print(auth_token)
            payment_link = create_payment_link(auth_token,payment_type.url,payment_type.login,payment_type.password,order)
            if payment_link['success']:
                print(payment_link['data'])
                order.payment_link = payment_link['data']['invoice_url']
                order.invoice_id = payment_link['data']['invoice_id']
                order.save()
                result = {"success": True}
            else:
                result = {"success": False, "message": 'Ошибка получения платежной ссылки'}
        else:
            result = {"success":False, "message":'Ошибка получения доступа к paykeeper'}
        return Response(result,status=200)

class Fill(APIView):
    def get(self,request):
        from openpyxl import load_workbook
        orders = Order.objects.all()
        orders.delete()
        wb = load_workbook(filename='ordes.xlsx')
        sheet_obj = wb.active
        max_row = sheet_obj.max_row

        # Loop will print all columns name max_row + 1

        for i in range(2, max_row + 1):
            old_id = sheet_obj.cell(row=i, column=1)
            created = sheet_obj.cell(row=i, column=3)
            address = sheet_obj.cell(row=i, column=9)
            comment = sheet_obj.cell(row=i, column=10)
            price = sheet_obj.cell(row=i, column=11)
            payer_old_id = sheet_obj.cell(row=i, column=18)
            #print(old_id.value,created.value,address.value,comment.value,price.value,payer_old_id.value)


            try:
                contractor = Contractor.objects.get(old_id=payer_old_id.value)
                client = Client.objects.get(id=contractor.client_id)
                print(client)
                Order.objects.create(
                    old_id=old_id.value,
                    client=client,
                    contractor=contractor,
                    delivery_address=address.value,
                    created_at_date=created.value,
                    orderComment=comment.value,
                    total_price=price.value,
                )
            except:
                pass


        return Response(status=200)
