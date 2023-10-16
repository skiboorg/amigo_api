import time
import datetime
from rest_framework.pagination import PageNumberPagination
from rest_framework.views import APIView
from selenium import webdriver
from selenium.webdriver.common.by import By

from .serializers import *
from .models import *
from rest_framework import generics, viewsets, parsers
from rest_framework.response import Response

import django_filters
from django.db.models import Q
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters

from data.models import City
from user.models import User

class ClientPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 10000


class ClientFilter(django_filters.FilterSet):
    q = django_filters.CharFilter(method='my_custom_filter', label="Search")
    def my_custom_filter(self, queryset, name, value):
        return queryset.filter(
            Q(fio__icontains=value)

        )
    class Meta:
        model = Client
        fields = ['fio','status__id','manager__id']

class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all()
    #serializer_class = ClientSerializer
    pagination_class = ClientPagination
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = ClientFilter

    def get_serializer_class(self):
        is_full = self.request.query_params.get('full',None)
        if is_full:
            return ClientSerializer
        else:
            return ClientForTableSerializer

    def create(self, request, *args, **kwargs):
        print(request.data)
        data = request.data
        new_client = Client.objects.create(
            manager_id= data.get('manager',None),
            status_id = data.get('status',None),
            city_id = data.get('city',None),
            fio = data.get('fio',None),
            web = data.get('web',None),
            address = data.get('address',None),
            comment = data.get('comment',None),
            created_at=datetime.datetime.now().date()
        )
        new_client.category.add(*data.get('category',None))

        return Response(status=status.HTTP_201_CREATED)

class ContactViewSet(viewsets.ModelViewSet):
    serializer_class = ContactSerializer
    queryset = Contact.objects.all()

class NoteViewSet(viewsets.ModelViewSet):
    serializer_class = NoteSerializer
    queryset = Note.objects.all()

class ContractorViewSet(viewsets.ModelViewSet):
    serializer_class = ContractorSerializer
    queryset = Contractor.objects.all()

class AddressViewSet(viewsets.ModelViewSet):
    serializer_class = DeliveryAddressSerializer
    queryset = DeliveryAddress.objects.all()

    # def update(self, request, *args, **kwargs):
    #     print(request.data)
class GetCategory(generics.ListAPIView):
    serializer_class = CategorySerializer
    queryset = Category.objects.all()

class GetStatus(generics.ListAPIView):
    serializer_class = StatusSerializer
    queryset = Status.objects.all()


class Fill(APIView):
    def get(self,request):
        from openpyxl import load_workbook
        clients = Client.objects.all()
        clients.delete()
        wb = load_workbook(filename='clients.xlsx')
        sheet_obj = wb.active
        max_row = sheet_obj.max_row

        # Loop will print all columns name

        for i in range(2, max_row + 1):

            old_id = sheet_obj.cell(row=i, column=1)
            fio = sheet_obj.cell(row=i, column=2)
            created_at = sheet_obj.cell(row=i, column=3)
            status_id = sheet_obj.cell(row=i, column=4)
            address = sheet_obj.cell(row=i, column=5)
            web = sheet_obj.cell(row=i, column=7)
            magager_id = sheet_obj.cell(row=i, column=8)
            comment = sheet_obj.cell(row=i, column=9)
            print(old_id.value,fio.value,created_at.value,status_id.value,address.value,web.value,magager_id.value,comment.value)
            try:
                magager = User.objects.get(old_id=magager_id.value)
                Client.objects.create(
                    old_id=old_id.value,
                    manager=magager,
                    status_id=status_id.value,
                    fio=fio.value,
                    web=web.value,
                    address=address.value,
                    comment=comment.value,
                    created_at=created_at.value,
                )
            except:
                pass


        return Response(status=200)
class FillContactor(APIView):
    def get(self,request):
        from openpyxl import load_workbook
        contractors = Contractor.objects.all()
        contractors.delete()
        wb = load_workbook(filename='payers.xlsx')
        sheet_obj = wb.active
        max_row = sheet_obj.max_row

        # Loop will print all columns name

        for i in range(2, max_row + 1):
            old_id = sheet_obj.cell(row=i, column=2)
            name = sheet_obj.cell(row=i, column=4)
            inn = sheet_obj.cell(row=i, column=5)
            client_old_id = sheet_obj.cell(row=i, column=6)
            print(old_id.value,name.value,inn.value,client_old_id.value)
            try:
                client = Client.objects.get(old_id=client_old_id.value)
                Contractor.objects.create(
                    old_id=old_id.value,
                    client=client,
                    name=name.value,
                    inn=inn.value,
                )
            except:
                pass


        return Response(status=200)